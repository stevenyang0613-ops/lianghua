"""西部量化可转债策略 V3.0 回测可视化模块

功能:
- 收益曲线图
- 回撤分析图
- 持仓分布图
- 月度收益热力图
- PDF报告导出
- Excel报告导出
"""
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import List, Dict, Optional, Any
from enum import Enum
import logging
import json
import io
import base64

logger = logging.getLogger(__name__)

# 检查可视化库
try:
    import matplotlib
    matplotlib.use('Agg')  # 非交互式后端
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.gridspec import GridSpec
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


# ============ 枚举类型 ============

class ChartType(str, Enum):
    """图表类型"""
    LINE = "line"
    AREA = "area"
    BAR = "bar"
    PIE = "pie"
    HEATMAP = "heatmap"
    SCATTER = "scatter"


class ReportFormat(str, Enum):
    """报告格式"""
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"
    JSON = "json"


# ============ 数据模型 ============

@dataclass
class BacktestChartData:
    """回测图表数据"""
    dates: List[str]
    nav_curve: List[float]
    benchmark_curve: List[float] = None
    drawdown_curve: List[float] = None
    returns: List[float] = None
    positions: Dict[str, List[float]] = None
    trades: List[Dict] = None

    def to_dict(self) -> dict:
        return {
            "dates": self.dates,
            "nav_curve": self.nav_curve,
            "benchmark_curve": self.benchmark_curve,
            "drawdown_curve": self.drawdown_curve,
            "returns": self.returns,
        }


@dataclass
class BacktestMetrics:
    """回测指标"""
    total_return: float
    annualized_return: float
    max_drawdown: float
    sharpe_ratio: float
    sortino_ratio: float
    calmar_ratio: float
    win_rate: float
    profit_factor: float
    avg_holding_days: float
    total_trades: int
    winning_trades: int
    losing_trades: int

    def to_dict(self) -> dict:
        return {
            "total_return": round(self.total_return * 100, 2),
            "annualized_return": round(self.annualized_return * 100, 2),
            "max_drawdown": round(self.max_drawdown * 100, 2),
            "sharpe_ratio": round(self.sharpe_ratio, 2),
            "sortino_ratio": round(self.sortino_ratio, 2),
            "calmar_ratio": round(self.calmar_ratio, 2),
            "win_rate": round(self.win_rate * 100, 2),
            "profit_factor": round(self.profit_factor, 2),
            "avg_holding_days": round(self.avg_holding_days, 1),
            "total_trades": self.total_trades,
            "winning_trades": self.winning_trades,
            "losing_trades": self.losing_trades,
        }


# ============ 图表生成器 ============

class ChartGenerator:
    """图表生成器"""

    def __init__(self, style: str = "seaborn-v0_8-whitegrid"):
        self.style = style
        self._setup_style()

    def _setup_style(self):
        """设置样式"""
        if MATPLOTLIB_AVAILABLE:
            try:
                plt.style.use(self.style)
            except Exception as e:
                logger.debug(f"[Visualization] style '{self.style}' not found: {e}")
                plt.style.use('seaborn-v0_8-whitegrid')
            plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False

    def generate_nav_curve(
        self,
        dates: List[str],
        nav_curve: List[float],
        benchmark_curve: List[float] = None,
        title: str = "净值曲线",
    ) -> Optional[str]:
        """生成净值曲线图"""
        if not MATPLOTLIB_AVAILABLE:
            return None

        fig, ax = plt.subplots(figsize=(12, 6))

        # 转换日期
        date_objs = [datetime.strptime(d, "%Y-%m-%d") for d in dates]

        # 绘制净值曲线
        ax.plot(date_objs, nav_curve, label="策略净值", linewidth=2, color='#2E86AB')

        if benchmark_curve:
            ax.plot(date_objs, benchmark_curve, label="基准", linewidth=1.5, color='#A23B72', linestyle='--')

        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xlabel("日期", fontsize=12)
        ax.set_ylabel("净值", fontsize=12)
        ax.legend(loc='upper left', fontsize=10)
        ax.grid(True, alpha=0.3)

        # 格式化x轴日期
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        plt.xticks(rotation=45)

        plt.tight_layout()
        return self._fig_to_base64(fig)

    def generate_drawdown_chart(
        self,
        dates: List[str],
        drawdown_curve: List[float],
        title: str = "回撤分析",
    ) -> Optional[str]:
        """生成回撤图"""
        if not MATPLOTLIB_AVAILABLE:
            return None

        fig, ax = plt.subplots(figsize=(12, 4))

        date_objs = [datetime.strptime(d, "%Y-%m-%d") for d in dates]

        # 填充区域
        ax.fill_between(date_objs, drawdown_curve, 0, alpha=0.5, color='#E74C3C', label="回撤")
        ax.plot(date_objs, drawdown_curve, color='#C0392B', linewidth=1)

        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xlabel("日期", fontsize=12)
        ax.set_ylabel("回撤率", fontsize=12)
        ax.legend(loc='lower left', fontsize=10)
        ax.grid(True, alpha=0.3)

        # 设置y轴为百分比
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y*100:.0f}%'))

        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        plt.xticks(rotation=45)

        plt.tight_layout()
        return self._fig_to_base64(fig)

    def generate_monthly_returns_heatmap(
        self,
        monthly_returns: Dict[str, Dict[str, float]],
        title: str = "月度收益热力图",
    ) -> Optional[str]:
        """生成月度收益热力图"""
        if not MATPLOTLIB_AVAILABLE:
            return None

        # 转换数据
        years = sorted(set(int(y) for y in monthly_returns.keys()))
        months = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']

        data = []
        for year in years:
            row = []
            year_data = monthly_returns.get(str(year), {})
            for m in range(1, 13):
                row.append(year_data.get(str(m), 0) * 100)
            data.append(row)

        fig, ax = plt.subplots(figsize=(14, len(years) * 0.8 + 2))

        # 绘制热力图
        sns.heatmap(
            data,
            annot=True,
            fmt='.1f',
            cmap='RdYlGn',
            center=0,
            xticklabels=months,
            yticklabels=years,
            ax=ax,
            cbar_kws={'label': '收益率(%)'},
        )

        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xlabel("月份", fontsize=12)
        ax.set_ylabel("年份", fontsize=12)

        plt.tight_layout()
        return self._fig_to_base64(fig)

    def generate_position_pie_chart(
        self,
        positions: Dict[str, float],
        title: str = "持仓分布",
    ) -> Optional[str]:
        """生成持仓饼图"""
        if not MATPLOTLIB_AVAILABLE:
            return None

        fig, ax = plt.subplots(figsize=(10, 8))

        # 排序
        sorted_positions = sorted(positions.items(), key=lambda x: x[1], reverse=True)
        labels = [p[0] for p in sorted_positions]
        sizes = [p[1] for p in sorted_positions]

        # 只显示前10个，其余归类为其他
        if len(labels) > 10:
            labels = labels[:10] + ['其他']
            sizes = sizes[:10] + [sum(sizes[10:])]

        colors = plt.cm.Set3(range(len(labels)))

        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%',
            colors=colors,
            startangle=90,
            pctdistance=0.85,
        )

        for autotext in autotexts:
            autotext.set_fontsize(9)

        ax.set_title(title, fontsize=14, fontweight='bold')

        plt.tight_layout()
        return self._fig_to_base64(fig)

    def generate_returns_distribution(
        self,
        returns: List[float],
        title: str = "收益分布",
    ) -> Optional[str]:
        """生成收益分布图"""
        if not MATPLOTLIB_AVAILABLE:
            return None

        fig, ax = plt.subplots(figsize=(10, 6))

        # 绘制直方图
        n, bins, patches = ax.hist(
            returns,
            bins=50,
            density=True,
            alpha=0.7,
            color='#3498DB',
            edgecolor='white',
        )

        # 添加正态分布曲线
        import numpy as np
        mu, sigma = np.mean(returns), np.std(returns)
        x = np.linspace(min(returns), max(returns), 100)
        ax.plot(x, 1/(sigma * np.sqrt(2 * np.pi)) * np.exp(-(x - mu)**2 / (2 * sigma**2)),
                linewidth=2, color='#E74C3C', label='正态分布')

        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xlabel("收益率", fontsize=12)
        ax.set_ylabel("频率", fontsize=12)
        ax.legend(fontsize=10)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        return self._fig_to_base64(fig)

    def generate_trades_scatter(
        self,
        trades: List[Dict],
        title: str = "交易分析",
    ) -> Optional[str]:
        """生成交易散点图"""
        if not MATPLOTLIB_AVAILABLE or not trades:
            return None

        fig, ax = plt.subplots(figsize=(10, 8))

        # 提取数据
        holding_days = [t.get('holding_days', 0) for t in trades]
        returns = [t.get('return', 0) * 100 for t in trades]
        colors = ['#27AE60' if r > 0 else '#E74C3C' for r in returns]

        ax.scatter(holding_days, returns, c=colors, alpha=0.6, s=50)

        ax.axhline(y=0, color='gray', linestyle='--', alpha=0.5)

        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xlabel("持有天数", fontsize=12)
        ax.set_ylabel("收益率(%)", fontsize=12)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        return self._fig_to_base64(fig)

    def _fig_to_base64(self, fig) -> str:
        """将图表转换为Base64"""
        buffer = io.BytesIO()
        fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
        plt.close(fig)
        buffer.seek(0)
        return base64.b64encode(buffer.getvalue()).decode()


# ============ 报告生成器 ============

class ReportGenerator:
    """报告生成器"""

    def __init__(self):
        self.chart_generator = ChartGenerator()

    def generate_html_report(
        self,
        metrics: BacktestMetrics,
        chart_data: BacktestChartData,
        strategy_name: str = "西部量化可转债策略",
    ) -> str:
        """生成HTML报告"""
        # 生成图表
        nav_chart = self.chart_generator.generate_nav_curve(
            chart_data.dates,
            chart_data.nav_curve,
            chart_data.benchmark_curve,
        )

        drawdown_chart = self.chart_generator.generate_drawdown_chart(
            chart_data.dates,
            chart_data.drawdown_curve,
        )

        returns_chart = None
        if chart_data.returns:
            returns_chart = self.chart_generator.generate_returns_distribution(chart_data.returns)

        # 构建HTML
        html = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{strategy_name} 回测报告</title>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2E86AB;
            text-align: center;
            border-bottom: 2px solid #2E86AB;
            padding-bottom: 10px;
        }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 15px;
            margin: 20px 0;
        }}
        .metric-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 15px;
            border-radius: 8px;
            text-align: center;
        }}
        .metric-value {{
            font-size: 24px;
            font-weight: bold;
            margin: 10px 0;
        }}
        .metric-label {{
            font-size: 12px;
            opacity: 0.9;
        }}
        .positive {{ color: #27AE60; }}
        .negative {{ color: #E74C3C; }}
        .chart-section {{
            margin: 30px 0;
        }}
        .chart-title {{
            font-size: 16px;
            font-weight: bold;
            color: #333;
            margin-bottom: 10px;
            padding-left: 10px;
            border-left: 4px solid #2E86AB;
        }}
        .chart-container {{
            text-align: center;
            background: #fafafa;
            padding: 20px;
            border-radius: 8px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }}
        th {{
            background-color: #2E86AB;
            color: white;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{strategy_name} 回测报告</h1>

        <div class="metrics-grid">
            <div class="metric-card">
                <div class="metric-label">总收益率</div>
                <div class="metric-value {'positive' if metrics.total_return > 0 else 'negative'}">{metrics.total_return*100:.2f}%</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">年化收益</div>
                <div class="metric-value {'positive' if metrics.annualized_return > 0 else 'negative'}">{metrics.annualized_return*100:.2f}%</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">最大回撤</div>
                <div class="metric-value negative">{metrics.max_drawdown*100:.2f}%</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">夏普比率</div>
                <div class="metric-value">{metrics.sharpe_ratio:.2f}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">胜率</div>
                <div class="metric-value">{metrics.win_rate*100:.1f}%</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">盈亏比</div>
                <div class="metric-value">{metrics.profit_factor:.2f}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">交易次数</div>
                <div class="metric-value">{metrics.total_trades}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">平均持仓</div>
                <div class="metric-value">{metrics.avg_holding_days:.1f}天</div>
            </div>
        </div>
        """

        if nav_chart:
            html += f"""
        <div class="chart-section">
            <div class="chart-title">净值曲线</div>
            <div class="chart-container">
                <img src="data:image/png;base64,{nav_chart}" alt="净值曲线">
            </div>
        </div>
            """

        if drawdown_chart:
            html += f"""
        <div class="chart-section">
            <div class="chart-title">回撤分析</div>
            <div class="chart-container">
                <img src="data:image/png;base64,{drawdown_chart}" alt="回撤分析">
            </div>
        </div>
            """

        if returns_chart:
            html += f"""
        <div class="chart-section">
            <div class="chart-title">收益分布</div>
            <div class="chart-container">
                <img src="data:image/png;base64,{returns_chart}" alt="收益分布">
            </div>
        </div>
            """

        html += f"""
        <div class="footer">
            <p>报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>西部量化可转债策略 V3.0</p>
        </div>
    </div>
</body>
</html>
        """

        return html

    def generate_excel_report(
        self,
        metrics: BacktestMetrics,
        chart_data: BacktestChartData,
        trades: List[Dict] = None,
        strategy_name: str = "西部量化可转债策略",
    ) -> Optional[bytes]:
        """生成Excel报告"""
        if not EXCEL_AVAILABLE:
            return None

        wb = Workbook()

        # 样式定义
        title_font = Font(bold=True, size=14, color="FFFFFF")
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # 概览页
        ws = wb.active
        ws.title = "概览"

        ws['A1'] = f"{strategy_name} 回测报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # 指标表
        metrics_data = [
            ["指标", "数值"],
            ["总收益率", f"{metrics.total_return*100:.2f}%"],
            ["年化收益", f"{metrics.annualized_return*100:.2f}%"],
            ["最大回撤", f"{metrics.max_drawdown*100:.2f}%"],
            ["夏普比率", f"{metrics.sharpe_ratio:.2f}"],
            ["索提诺比率", f"{metrics.sortino_ratio:.2f}"],
            ["卡玛比率", f"{metrics.calmar_ratio:.2f}"],
            ["胜率", f"{metrics.win_rate*100:.1f}%"],
            ["盈亏比", f"{metrics.profit_factor:.2f}"],
            ["交易次数", metrics.total_trades],
            ["盈利次数", metrics.winning_trades],
            ["亏损次数", metrics.losing_trades],
            ["平均持仓天数", f"{metrics.avg_holding_days:.1f}"],
        ]

        for row_idx, row_data in enumerate(metrics_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF")

        # 净值数据页
        ws_nav = wb.create_sheet("净值数据")

        nav_headers = ["日期", "净值", "回撤"]
        for col_idx, header in enumerate(nav_headers, start=1):
            cell = ws_nav.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = border

        for row_idx, (date, nav, dd) in enumerate(
            zip(chart_data.dates, chart_data.nav_curve, chart_data.drawdown_curve), start=2
        ):
            ws_nav.cell(row=row_idx, column=1, value=date).border = border
            ws_nav.cell(row=row_idx, column=2, value=nav).border = border
            ws_nav.cell(row=row_idx, column=3, value=dd).border = border

        # 交易记录页
        if trades:
            ws_trades = wb.create_sheet("交易记录")

            trade_headers = ["日期", "代码", "名称", "操作", "数量", "价格", "盈亏"]
            for col_idx, header in enumerate(trade_headers, start=1):
                cell = ws_trades.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.border = border

            for row_idx, trade in enumerate(trades, start=2):
                ws_trades.cell(row=row_idx, column=1, value=trade.get('date')).border = border
                ws_trades.cell(row=row_idx, column=2, value=trade.get('code')).border = border
                ws_trades.cell(row=row_idx, column=3, value=trade.get('name')).border = border
                ws_trades.cell(row=row_idx, column=4, value=trade.get('action')).border = border
                ws_trades.cell(row=row_idx, column=5, value=trade.get('quantity')).border = border
                ws_trades.cell(row=row_idx, column=6, value=trade.get('price')).border = border
                ws_trades.cell(row=row_idx, column=7, value=trade.get('profit')).border = border

        # 导出
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_pdf_report(
        self,
        metrics: BacktestMetrics,
        chart_data: BacktestChartData,
        strategy_name: str = "西部量化可转债策略",
    ) -> Optional[bytes]:
        """生成PDF报告"""
        if not PDF_AVAILABLE:
            return None

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # 居中
        )

        elements = []

        # 标题
        elements.append(Paragraph(f"{strategy_name} 回测报告", title_style))
        elements.append(Spacer(1, 20))

        # 指标表
        metrics_table_data = [
            ["指标", "数值"],
            ["总收益率", f"{metrics.total_return*100:.2f}%"],
            ["年化收益", f"{metrics.annualized_return*100:.2f}%"],
            ["最大回撤", f"{metrics.max_drawdown*100:.2f}%"],
            ["夏普比率", f"{metrics.sharpe_ratio:.2f}"],
            ["胜率", f"{metrics.win_rate*100:.1f}%"],
            ["盈亏比", f"{metrics.profit_factor:.2f}"],
            ["交易次数", str(metrics.total_trades)],
        ]

        metrics_table = Table(metrics_table_data, colWidths=[5*cm, 5*cm])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(metrics_table)
        elements.append(Spacer(1, 20))

        # 图表
        nav_chart = self.chart_generator.generate_nav_curve(
            chart_data.dates,
            chart_data.nav_curve,
            chart_data.benchmark_curve,
        )

        if nav_chart:
            img_data = base64.b64decode(nav_chart)
            img = Image(io.BytesIO(img_data), width=15*cm, height=7.5*cm)
            elements.append(img)
            elements.append(Spacer(1, 10))

        drawdown_chart = self.chart_generator.generate_drawdown_chart(
            chart_data.dates,
            chart_data.drawdown_curve,
        )

        if drawdown_chart:
            img_data = base64.b64decode(drawdown_chart)
            img = Image(io.BytesIO(img_data), width=15*cm, height=5*cm)
            elements.append(img)

        # 页脚
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal'],
        ))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


# ============ 可视化管理器 ============

class BacktestVisualizer:
    """回测可视化统一接口"""

    def __init__(self):
        self.chart_generator = ChartGenerator()
        self.report_generator = ReportGenerator()

    def generate_all_charts(
        self,
        chart_data: BacktestChartData,
    ) -> Dict[str, str]:
        """生成所有图表"""
        charts = {}

        # 净值曲线
        charts['nav_curve'] = self.chart_generator.generate_nav_curve(
            chart_data.dates,
            chart_data.nav_curve,
            chart_data.benchmark_curve,
        )

        # 回撤图
        if chart_data.drawdown_curve:
            charts['drawdown'] = self.chart_generator.generate_drawdown_chart(
                chart_data.dates,
                chart_data.drawdown_curve,
            )

        # 收益分布
        if chart_data.returns:
            charts['returns_distribution'] = self.chart_generator.generate_returns_distribution(
                chart_data.returns,
            )

        # 持仓分布
        if chart_data.positions:
            position_weights = {}
            for code, weights in chart_data.positions.items():
                position_weights[code] = weights[-1] if weights else 0
            charts['position_pie'] = self.chart_generator.generate_position_pie_chart(
                position_weights,
            )

        # 交易分析
        if chart_data.trades:
            charts['trades_scatter'] = self.chart_generator.generate_trades_scatter(
                chart_data.trades,
            )

        return charts

    def generate_report(
        self,
        metrics: BacktestMetrics,
        chart_data: BacktestChartData,
        format: ReportFormat = ReportFormat.HTML,
        strategy_name: str = "西部量化可转债策略",
    ) -> Any:
        """生成报告"""
        if format == ReportFormat.HTML:
            return self.report_generator.generate_html_report(metrics, chart_data, strategy_name)
        elif format == ReportFormat.EXCEL:
            return self.report_generator.generate_excel_report(metrics, chart_data, None, strategy_name)
        elif format == ReportFormat.PDF:
            return self.report_generator.generate_pdf_report(metrics, chart_data, strategy_name)
        elif format == ReportFormat.JSON:
            return json.dumps({
                "metrics": metrics.to_dict(),
                "chart_data": chart_data.to_dict(),
            }, ensure_ascii=False, indent=2)

        return None


# ============ 便捷函数 ============

def plot_backtest(
    dates: List[str],
    nav_curve: List[float],
    benchmark_curve: List[float] = None,
) -> Optional[str]:
    """绘制回测净值曲线"""
    generator = ChartGenerator()
    return generator.generate_nav_curve(dates, nav_curve, benchmark_curve)


def generate_backtest_report(
    metrics: Dict[str, float],
    chart_data: Dict[str, Any],
    format: str = "html",
) -> Any:
    """生成回测报告"""
    visualizer = BacktestVisualizer()

    metrics_obj = BacktestMetrics(
        total_return=metrics.get('total_return', 0),
        annualized_return=metrics.get('annualized_return', 0),
        max_drawdown=metrics.get('max_drawdown', 0),
        sharpe_ratio=metrics.get('sharpe_ratio', 0),
        sortino_ratio=metrics.get('sortino_ratio', 0),
        calmar_ratio=metrics.get('calmar_ratio', 0),
        win_rate=metrics.get('win_rate', 0),
        profit_factor=metrics.get('profit_factor', 0),
        avg_holding_days=metrics.get('avg_holding_days', 0),
        total_trades=metrics.get('total_trades', 0),
        winning_trades=metrics.get('winning_trades', 0),
        losing_trades=metrics.get('losing_trades', 0),
    )

    chart_obj = BacktestChartData(
        dates=chart_data.get('dates', []),
        nav_curve=chart_data.get('nav_curve', []),
        benchmark_curve=chart_data.get('benchmark_curve'),
        drawdown_curve=chart_data.get('drawdown_curve'),
        returns=chart_data.get('returns'),
    )

    report_format = ReportFormat(format)

    return visualizer.generate_report(metrics_obj, chart_obj, report_format)
